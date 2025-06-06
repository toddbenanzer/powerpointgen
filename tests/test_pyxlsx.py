import unittest
import pandas as pd
from pyxlsx import PyWorkbook, CHART_TYPE_BAR, CHART_TYPE_COLUMN, CHART_TYPE_LINE, CHART_TYPE_PIE
import os
from openpyxl import load_workbook # For direct verification
from openpyxl.chart import BarChart, LineChart, PieChart # For isinstance checks

# Define a directory for test outputs
TEST_OUTPUT_DIR = "test_outputs"
if not os.path.exists(TEST_OUTPUT_DIR):
    os.makedirs(TEST_OUTPUT_DIR)

class TestPyXLSX(unittest.TestCase):

    def test_create_workbook_and_add_sheet(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_create_workbook.xlsx")

        wb = PyWorkbook()
        self.assertIsNotNone(wb.workbook, "Workbook object should be created.")

        # Test adding sheet by title
        ws1 = wb.add_worksheet(title="MySheet1")
        self.assertEqual(ws1.worksheet.title, "MySheet1", "Sheet title not set correctly.")
        self.assertIn("MySheet1", wb.workbook.sheetnames, "Sheet not found in workbook by title.")

        # Test adding sheet with default title
        ws2 = wb.add_worksheet()
        self.assertIsNotNone(ws2.worksheet.title, "Default sheet title should be assigned.")
        self.assertIn(ws2.worksheet.title, wb.workbook.sheetnames, "Default titled sheet not found.")

        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path), "Workbook file not saved.")

        # Clean up
        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_write_simple_dataframe(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_write_dataframe.xlsx")
        data = {'col1': [1, 2], 'col2': ['A', 'B']}
        df = pd.DataFrame(data)

        wb = PyWorkbook()
        ws = wb.add_worksheet(title="DataFrameSheet")
        ws.write_dataframe(df, start_row=1, start_col=1, header=True)
        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        # Verification using openpyxl directly
        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["DataFrameSheet"]

        self.assertEqual(verify_ws.cell(row=1, column=1).value, 'col1')
        self.assertEqual(verify_ws.cell(row=1, column=2).value, 'col2')
        self.assertEqual(verify_ws.cell(row=2, column=1).value, 1)
        self.assertEqual(verify_ws.cell(row=2, column=2).value, 'A')
        self.assertEqual(verify_ws.cell(row=3, column=1).value, 2)
        self.assertEqual(verify_ws.cell(row=3, column=2).value, 'B')

        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_write_dataframe_with_formatting(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_dataframe_formatting.xlsx")
        data = {'ID': [101, 102], 'Value': [123.45, 678.90], 'Status': ['Active', 'Inactive']}
        df = pd.DataFrame(data)

        wb = PyWorkbook()
        ws = wb.add_worksheet("FormattedDF")
        ws.write_dataframe(df, start_row=2, start_col=2, header=True, include_index=False,
                           header_font_bold=True, header_fill_color="C0C0C0", # Silver
                           font_name="Calibri", font_size=11,
                           number_formats={'Value': '0.00'},
                           column_widths={'B': 15, 'C': 20, 'D': 10}, # Test with column letters
                           data_alignment_horizontal='center',
                           header_alignment_horizontal='center'
                           )
        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        # Basic verification (more detailed style checking can be complex and added later)
        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["FormattedDF"]

        # Header check
        self.assertEqual(verify_ws.cell(row=2, column=2).value, 'ID')
        self.assertTrue(verify_ws.cell(row=2, column=2).font.bold, "Header font should be bold.")
        self.assertIsNotNone(verify_ws.cell(row=2, column=2).fill.fgColor.rgb, "Header fill color not applied.")


        # Data check
        self.assertEqual(verify_ws.cell(row=3, column=3).value, 123.45)
        self.assertEqual(verify_ws.cell(row=3, column=3).number_format, '0.00', "Number format not applied.")
        self.assertEqual(verify_ws.cell(row=3, column=3).font.name, 'Calibri', "Data font name not applied.")
        self.assertEqual(verify_ws.cell(row=3, column=3).alignment.horizontal, 'center', "Data alignment not applied.")


        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_write_and_read_cell(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_cell_ops.xlsx")
        wb = PyWorkbook()
        ws = wb.add_worksheet("CellOps")

        ws.write_cell(1, 1, "Hello", font_bold=True, font_color="FF0000") # Red text
        ws.write_cell(2, 1, 12345, number_format="#,##0")

        self.assertEqual(ws.read_cell(1, 1), "Hello")
        self.assertEqual(ws.read_cell(2, 1), 12345)

        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        # Verify with openpyxl
        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["CellOps"]
        self.assertEqual(verify_ws.cell(1,1).value, "Hello")
        self.assertTrue(verify_ws.cell(1,1).font.bold)
        # self.assertEqual(verify_ws.cell(1,1).font.color.rgb, "FFFF0000") # Note: openpyxl might add FF for alpha

        self.assertEqual(verify_ws.cell(2,1).value, 12345)
        self.assertEqual(verify_ws.cell(2,1).number_format, "#,##0")

        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_add_bar_chart(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_bar_chart.xlsx")
        wb = PyWorkbook()
        ws = wb.add_worksheet("BarChartSheet")

        # Sample Data
        ws.write_cell(1, 1, "Category")
        ws.write_cell(2, 1, "Cat A")
        ws.write_cell(3, 1, "Cat B")
        ws.write_cell(4, 1, "Cat C")

        ws.write_cell(1, 2, "Series 1")
        ws.write_cell(2, 2, 10)
        ws.write_cell(3, 2, 20)
        ws.write_cell(4, 2, 30)

        ws.write_cell(1, 3, "Series 2")
        ws.write_cell(2, 3, 15)
        ws.write_cell(3, 3, 25)
        ws.write_cell(4, 3, 35)

        chart_title = "My Bar Chart"
        ws.add_chart(
            chart_type=CHART_TYPE_BAR,
            cell_anchor="E2",
            series_data_range=(1, 2, 4, 3),
            category_labels_range=(2, 1, 4, 1),
            title=chart_title,
            x_axis_label="Categories",
            y_axis_label="Values"
        )
        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["BarChartSheet"]
        self.assertEqual(len(verify_ws._charts), 1)
        chart = verify_ws._charts[0]
        self.assertIsInstance(chart, BarChart)
        self.assertEqual(chart.title.tx.rich.p[0].r[0].t, chart_title) # More robust title check
        self.assertEqual(chart.x_axis.title.tx.rich.p[0].r[0].t, "Categories")
        self.assertEqual(chart.y_axis.title.tx.rich.p[0].r[0].t, "Values")
        self.assertEqual(chart.type, "bar")

        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_add_column_chart(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_column_chart.xlsx")
        wb = PyWorkbook()
        ws = wb.add_worksheet("ColumnChartSheet")

        ws.write_cell(1, 1, "Period")
        ws.write_cell(2, 1, "Q1")
        ws.write_cell(3, 1, "Q2")
        ws.write_cell(4, 1, "Q3")

        ws.write_cell(1, 2, "Product X")
        ws.write_cell(2, 2, 50)
        ws.write_cell(3, 2, 70)
        ws.write_cell(4, 2, 60)

        chart_title = "My Column Chart"
        ws.add_chart(
            chart_type=CHART_TYPE_COLUMN,
            cell_anchor="D2",
            series_data_range=(1, 2, 4, 2),
            category_labels_range=(2, 1, 4, 1),
            title=chart_title,
            x_axis_label="Time Period",
            y_axis_label="Units Sold"
        )
        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["ColumnChartSheet"]
        self.assertEqual(len(verify_ws._charts), 1)
        chart = verify_ws._charts[0]
        self.assertIsInstance(chart, BarChart) # Column chart is a type of BarChart in openpyxl
        self.assertEqual(chart.title.tx.rich.p[0].r[0].t, chart_title)
        self.assertEqual(chart.x_axis.title.tx.rich.p[0].r[0].t, "Time Period")
        self.assertEqual(chart.y_axis.title.tx.rich.p[0].r[0].t, "Units Sold")
        self.assertEqual(chart.type, "col")


        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_add_line_chart(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_line_chart.xlsx")
        wb = PyWorkbook()
        ws = wb.add_worksheet("LineChartSheet")

        ws.write_cell(1, 1, "Month")
        ws.write_cell(2, 1, "Jan")
        ws.write_cell(3, 1, "Feb")
        ws.write_cell(4, 1, "Mar")

        ws.write_cell(1, 2, "Trend 1")
        ws.write_cell(2, 2, 100)
        ws.write_cell(3, 2, 120)
        ws.write_cell(4, 2, 90)

        chart_title = "My Line Chart"
        ws.add_chart(
            chart_type=CHART_TYPE_LINE,
            cell_anchor="D2",
            series_data_range=(1, 2, 4, 2),
            category_labels_range=(2, 1, 4, 1),
            title=chart_title,
            x_axis_label="Timeline",
            y_axis_label="Metric"
        )
        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["LineChartSheet"]
        self.assertEqual(len(verify_ws._charts), 1)
        chart = verify_ws._charts[0]
        self.assertIsInstance(chart, LineChart)
        self.assertEqual(chart.title.tx.rich.p[0].r[0].t, chart_title)
        self.assertEqual(chart.x_axis.title.tx.rich.p[0].r[0].t, "Timeline")
        self.assertEqual(chart.y_axis.title.tx.rich.p[0].r[0].t, "Metric")

        if os.path.exists(wb_path):
            os.remove(wb_path)

    def test_add_pie_chart(self):
        wb_path = os.path.join(TEST_OUTPUT_DIR, "test_pie_chart.xlsx")
        wb = PyWorkbook()
        ws = wb.add_worksheet("PieChartSheet")

        ws.write_cell(1, 1, "Segment") # Category labels
        ws.write_cell(2, 1, "Alpha")
        ws.write_cell(3, 1, "Beta")
        ws.write_cell(4, 1, "Gamma")

        ws.write_cell(1, 2, "Market Share") # Series title
        ws.write_cell(2, 2, 40)
        ws.write_cell(3, 2, 35)
        ws.write_cell(4, 2, 25)

        chart_title = "Market Distribution"
        ws.add_chart(
            chart_type=CHART_TYPE_PIE,
            cell_anchor="D2",
            series_data_range=(1, 2, 4, 2), # Data in B1:B4 (incl. header)
            category_labels_range=(2, 1, 4, 1), # Labels in A2:A4
            title=chart_title
        )
        wb.save(wb_path)
        self.assertTrue(os.path.exists(wb_path))

        verify_wb = load_workbook(wb_path)
        verify_ws = verify_wb["PieChartSheet"]
        self.assertEqual(len(verify_ws._charts), 1)
        chart = verify_ws._charts[0]
        self.assertIsInstance(chart, PieChart)
        self.assertEqual(chart.title.tx.rich.p[0].r[0].t, chart_title)

        if os.path.exists(wb_path):
            os.remove(wb_path)

if __name__ == '__main__':
    unittest.main()
