from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.utils.exceptions import InvalidFileException
# Placeholder for PyWorksheet, will be created in a later step
from .worksheet import PyWorksheet

class PyWorkbook:
    def __init__(self, filepath=None):
        """
        Initializes a PyWorkbook.

        Args:
            filepath (str, optional): Path to an existing .xlsx file to open.
                                      If None, a new workbook is created.
                                      Currently, only creating new workbooks is focused on.
        """
        if filepath:
            # For now, let's keep the focus on creating new workbooks as per initial plan.
            # Opening existing files can be enhanced later.
            # try:
            #     self.workbook = openpyxl.load_workbook(filepath)
            # except InvalidFileException:
            #     raise ValueError(f"Invalid Excel file or file not found: {filepath}")
            # For now, raise an error if filepath is provided until open functionality is built
            raise NotImplementedError("Opening existing files is not yet implemented. Please initialize with no arguments to create a new workbook.")
        else:
            self.workbook = OpenpyxlWorkbook()
            # Remove the default sheet created by openpyxl, or activate it.
            # For a cleaner start, let's remove it. User will explicitly add sheets.
            if "Sheet" in self.workbook.sheetnames and len(self.workbook.sheetnames) == 1:
                self.workbook.remove(self.workbook.active)
        self.filepath = filepath if filepath else "new_workbook.xlsx"


    def add_worksheet(self, title=None):
        """
        Adds a new worksheet to the workbook.

        Args:
            title (str, optional): The title for the new worksheet.
                                   If None, openpyxl will assign a default title.

        Returns:
            PyWorksheet: A PyWorksheet instance wrapping the new sheet. (Actual return type after PyWorksheet is implemented)
        """
        if title:
            new_ws = self.workbook.create_sheet(title=title)
        else:
            new_ws = self.workbook.create_sheet()

        return PyWorksheet(new_ws) # This will be uncommented/added when PyWorksheet exists
        # return new_ws # Returning the raw openpyxl worksheet for now

    def get_worksheet(self, sheet_name_or_index):
        """
        Retrieves an existing worksheet by its name or index.

        Args:
            sheet_name_or_index (str or int): The name or 0-based index of the worksheet.

        Returns:
            PyWorksheet: A PyWorksheet instance. (Actual return type after PyWorksheet is implemented)

        Raises:
            KeyError: If sheet_name is not found.
            IndexError: If sheet_index is out of range.
        """
        if isinstance(sheet_name_or_index, str):
            if sheet_name_or_index in self.workbook.sheetnames:
                ws = self.workbook[sheet_name_or_index]
                return PyWorksheet(ws) # For later
                # return ws # For now
            else:
                raise KeyError(f"Worksheet with name '{sheet_name_or_index}' not found.")
        elif isinstance(sheet_name_or_index, int):
            try:
                ws = self.workbook.worksheets[sheet_name_or_index]
                return PyWorksheet(ws) # For later
                # return ws # For now
            except IndexError:
                raise IndexError(f"Worksheet index {sheet_name_or_index} is out of range.")
        else:
            raise TypeError("sheet_name_or_index must be a string or an integer.")

    @property
    def active_worksheet(self):
        """
        Gets the currently active worksheet.

        Returns:
            PyWorksheet: A PyWorksheet instance. (Actual return type after PyWorksheet is implemented)
        """
        return PyWorksheet(self.workbook.active) # For later
        # return self.workbook.active # For now

    def save(self, filename=None):
        """
        Saves the workbook to the given filename.

        Args:
            filename (str, optional): The path to save the file.
                                      If None, uses the filepath provided during initialization
                                      or 'new_workbook.xlsx' if none was given.
        """
        save_path = filename if filename else self.filepath
        if not save_path.endswith('.xlsx'):
            save_path += '.xlsx'

        try:
            self.workbook.save(save_path)
            print(f"Workbook saved to {save_path}")
        except Exception as e:
            # Consider more specific exception handling if needed
            raise IOError(f"Could not save workbook to {save_path}: {e}")
