import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.styles.numbers import BUILTIN_FORMATS as OPENPYXL_BUILTIN_NUMBER_FORMATS
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from .constants import CHART_TYPE_BAR, CHART_TYPE_COLUMN, CHART_TYPE_LINE, CHART_TYPE_PIE # New import

# Consider moving default style values to constants.py later if they become numerous
DEFAULT_FONT_NAME = "Arial"
DEFAULT_FONT_SIZE = 10
DEFAULT_HEADER_FONT_BOLD = True
DEFAULT_HEADER_FILL_COLOR = "D9D9D9" # Light grey

class PyWorksheet:
    def __init__(self, openpyxl_worksheet):
        """
        Initializes a PyWorksheet.

        Args:
            openpyxl_worksheet (openpyxl.worksheet.worksheet.Worksheet):
                The underlying openpyxl worksheet object.
        """
        self.worksheet = openpyxl_worksheet
        # self._number_formats_reverse_lookup = {v: k for k, v in OPENPYXL_BUILTIN_NUMBER_FORMATS.items()} # Removed as openpyxl expects string format codes


    def _apply_cell_style(self, cell,
                          font_name=None, font_size=None, font_bold=False, font_italic=False, font_color=None,
                          fill_color=None, number_format=None,
                          align_horizontal=None, align_vertical=None,
                          wrap_text=None):
        """Internal helper to apply styles to a single cell."""
        # Font
        cell.font = Font(name=font_name or cell.font.name or DEFAULT_FONT_NAME,
                         size=font_size or cell.font.size or DEFAULT_FONT_SIZE,
                         bold=font_bold or cell.font.bold, # Respect existing if not overridden
                         italic=font_italic or cell.font.italic,
                         color=font_color or cell.font.color)

        # Fill
        if fill_color: # Only apply if a color is specified
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Number Format
        if number_format:
            # openpyxl expects the format code string directly.
            cell.number_format = number_format


        # Alignment
        current_alignment = cell.alignment or Alignment() # Get current or new
        new_alignment_args = {}
        if align_horizontal: new_alignment_args['horizontal'] = align_horizontal
        if align_vertical: new_alignment_args['vertical'] = align_vertical
        if wrap_text is not None: new_alignment_args['wrap_text'] = wrap_text

        if new_alignment_args:
             cell.alignment = Alignment(**{**current_alignment.__dict__, **new_alignment_args})


    def write_dataframe(self, df, start_row=1, start_col=1, include_index=False, index_label=None, header=True,
                          number_formats=None, # dict: {col_name: format_str}
                          font_name=None, font_size=None, font_bold=False, font_italic=False, font_color=None, # For data
                          header_font_name=None, header_font_size=None, header_font_bold=None,
                          header_font_italic=False, header_font_color=None, header_fill_color=None,
                          column_widths=None, # dict: {col_letter_or_idx: width} or 'auto' for all
                          row_heights=None, # dict: {row_idx: height}
                          data_alignment_horizontal=None, data_alignment_vertical=None, data_wrap_text=None,
                          header_alignment_horizontal=None, header_alignment_vertical=None, header_wrap_text=None
                          ):
        """
        Writes a pandas DataFrame to the worksheet with formatting.

        Args:
            df (pd.DataFrame): The DataFrame to write.
            start_row (int): 1-based row for the top-left cell of the header (or data if header=False).
            start_col (int): 1-based column for the top-left cell of the header (or data if header=False).
            include_index (bool): Whether to write the DataFrame index.
            index_label (str): Custom label for the index column if include_index is True.
            header (bool): Whether to write the DataFrame header.
            number_formats (dict): Maps column names (or index_label if include_index) to openpyxl number format strings.
                                   Example: {'Sales': '#,##0.00', 'Date': 'yyyy-mm-dd'}
            font_name, font_size, font_bold, font_italic, font_color: Style for data cells.
            header_font_name, ..., header_fill_color: Style for header cells.
            column_widths (dict or str): Dict maps column letters (e.g., 'A') or 1-based indices to widths.
                                        If 'auto', attempts to auto-fit all written columns.
            row_heights (dict): Maps 1-based row indices to heights.
            data_alignment_horizontal/vertical/wrap_text: Alignment for data cells.
            header_alignment_horizontal/vertical/wrap_text: Alignment for header cells.
        """
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Input 'df' must be a pandas DataFrame.")

        current_row = start_row

        # Prepare effective column names (including index if applicable)
        effective_df = df
        if include_index:
            effective_df = df.reset_index()
            if index_label: # Rename the new index column
                effective_df.rename(columns={effective_df.columns[0]: index_label}, inplace=True)

        col_names = list(effective_df.columns)

        # Write Header
        if header:
            for c_idx, col_name in enumerate(col_names):
                cell = self.worksheet.cell(row=current_row, column=start_col + c_idx, value=col_name)
                self._apply_cell_style(cell,
                                       font_name=header_font_name or font_name, # Fallback to data font
                                       font_size=header_font_size or font_size,
                                       font_bold=header_font_bold if header_font_bold is not None else DEFAULT_HEADER_FONT_BOLD,
                                       font_italic=header_font_italic, # Defaults to False
                                       font_color=header_font_color, # Defaults to None (black)
                                       fill_color=header_fill_color or DEFAULT_HEADER_FILL_COLOR,
                                       align_horizontal=header_alignment_horizontal,
                                       align_vertical=header_alignment_vertical,
                                       wrap_text=header_wrap_text)
            current_row += 1

        # Write Data
        for r_idx, row_data in enumerate(effective_df.itertuples(index=False, name=None)):
            for c_idx, cell_value in enumerate(row_data):
                # Handle specific pandas types that openpyxl might not like directly
                if pd.isna(cell_value):
                    cell_value = None # Write as empty cell
                # Add more type conversions if necessary (e.g., pd.Timestamp to datetime.datetime)
                elif isinstance(cell_value, pd.Timestamp):
                     cell_value = cell_value.to_pydatetime()


                cell = self.worksheet.cell(row=current_row + r_idx, column=start_col + c_idx, value=cell_value)

                col_name_for_formatting = col_names[c_idx]
                current_number_format = None
                if number_formats and col_name_for_formatting in number_formats:
                    current_number_format = number_formats[col_name_for_formatting]

                self._apply_cell_style(cell,
                                       font_name=font_name, font_size=font_size, font_bold=font_bold,
                                       font_italic=font_italic, font_color=font_color,
                                       number_format=current_number_format,
                                       align_horizontal=data_alignment_horizontal,
                                       align_vertical=data_alignment_vertical,
                                       wrap_text=data_wrap_text)

        # Column Widths
        if column_widths:
            if column_widths == 'auto':
                for c_idx, col_name in enumerate(col_names):
                    column_letter = get_column_letter(start_col + c_idx)
                    max_length = 0
                    # Check header length
                    if header:
                         max_length = max(max_length, len(str(col_name)))
                    # Check data length
                    for i in range(len(effective_df[col_name])):
                        val_str = str(effective_df[col_name].iloc[i])
                        if val_str is not None:
                             max_length = max(max_length, len(val_str))
                    adjusted_width = (max_length + 2) * 1.2 # Basic auto-fit heuristic
                    self.worksheet.column_dimensions[column_letter].width = adjusted_width

            elif isinstance(column_widths, dict):
                for col_ref, width in column_widths.items():
                    col_letter = col_ref if isinstance(col_ref, str) else get_column_letter(col_ref)
                    self.worksheet.column_dimensions[col_letter].width = width

        # Row Heights (applied to all written data rows + header)
        if row_heights:
            header_rows = 1 if header else 0
            for r_offset in range(header_rows + len(df)): # Iterate through written rows
                actual_row_idx = start_row + r_offset
                if actual_row_idx in row_heights:
                    self.worksheet.row_dimensions[actual_row_idx].height = row_heights[actual_row_idx]

    def write_cell(self, row, col, value,
                   font_name=None, font_size=None, font_bold=False, font_italic=False, font_color=None,
                   fill_color=None, number_format=None,
                   align_horizontal=None, align_vertical=None, wrap_text=None):
        """
        Writes a value to a specific cell with optional styling.
        Args:
            row (int): 1-based row index.
            col (int): 1-based column index.
            value: The value to write to the cell.
            font_name, ..., wrap_text: Styling options similar to _apply_cell_style.
        """
        cell = self.worksheet.cell(row=row, column=col, value=value)
        self._apply_cell_style(cell,
                               font_name=font_name, font_size=font_size, font_bold=font_bold,
                               font_italic=font_italic, font_color=font_color,
                               fill_color=fill_color, number_format=number_format,
                               align_horizontal=align_horizontal, align_vertical=align_vertical,
                               wrap_text=wrap_text)
        return cell # Return the openpyxl cell object

    def read_cell(self, row, col):
        """
        Reads a value from a specific cell.
        Args:
            row (int): 1-based row index.
            col (int): 1-based column index.
        Returns:
            The value of the cell.
        """
        return self.worksheet.cell(row=row, column=col).value

    def merge_cells(self, start_row, start_col, end_row, end_col):
        """
        Merges a range of cells.
        Args:
            start_row (int): 1-based row index of the top-left cell.
            start_col (int): 1-based column index of the top-left cell.
            end_row (int): 1-based row index of the bottom-right cell.
            end_col (int): 1-based column index of the bottom-right cell.
        """
        self.worksheet.merge_cells(start_row=start_row, start_column=start_col,
                                   end_row=end_row, end_column=end_col)

    def unmerge_cells(self, start_row, start_col, end_row, end_col):
        """
        Unmerges a range of cells.
        Args:
            start_row (int): 1-based row index of the top-left cell.
            start_col (int): 1-based column index of the top-left cell.
            end_row (int): 1-based row index of the bottom-right cell.
            end_col (int): 1-based column index of the bottom-right cell.
        """
        self.worksheet.unmerge_cells(start_row=start_row, start_column=start_col,
                                     end_row=end_row, end_column=end_col)

    def add_chart(self,
                  chart_type: str,
                  cell_anchor: str,
                  series_data_range: tuple, # (min_row, min_col, max_row, max_col)
                  category_labels_range: tuple = None, # (min_row, min_col, max_row, max_col)
                  titles_from_data: bool = True,
                  title: str = None,
                  x_axis_label: str = None,
                  y_axis_label: str = None,
                  chart_width: float = 15,  # cm
                  chart_height: float = 7.5 # cm
                 ):
        # Initial input validation (can be expanded later)
        # Convert input chart_type to upper for case-insensitive comparison
        input_chart_type_upper = chart_type.upper()
        if input_chart_type_upper not in [CHART_TYPE_BAR, CHART_TYPE_COLUMN, CHART_TYPE_LINE, CHART_TYPE_PIE]:
            raise ValueError(f"Unsupported chart_type: {chart_type}. Supported types: BAR, COLUMN, LINE, PIE.")

        if not (isinstance(series_data_range, tuple) and len(series_data_range) == 4):
            raise ValueError("series_data_range must be a tuple of (min_row, min_col, max_row, max_col).")
        if category_labels_range and not (isinstance(category_labels_range, tuple) and len(category_labels_range) == 4):
            raise ValueError("category_labels_range must be a tuple of (min_row, min_col, max_row, max_col) if provided.")

        chart_obj = None

        if input_chart_type_upper == CHART_TYPE_BAR or input_chart_type_upper == CHART_TYPE_COLUMN:
            chart_obj = BarChart()
            if input_chart_type_upper == CHART_TYPE_BAR:
                chart_obj.type = "bar" # Horizontal bars
            else: # CHART_TYPE_COLUMN
                chart_obj.type = "col" # Vertical columns

            s_min_row, s_min_col, s_max_row, s_max_col = series_data_range
            series_ref = Reference(self.worksheet,
                                   min_col=s_min_col, min_row=s_min_row,
                                   max_col=s_max_col, max_row=s_max_row)
            chart_obj.add_data(series_ref, titles_from_data=titles_from_data)

            if category_labels_range:
                c_min_row, c_min_col, c_max_row, c_max_col = category_labels_range
                cat_ref = Reference(self.worksheet,
                                    min_col=c_min_col, min_row=c_min_row,
                                    max_col=c_max_col, max_row=c_max_row)
                chart_obj.set_categories(cat_ref)

            if title:
                chart_obj.title = title
            if x_axis_label:
                chart_obj.x_axis.title = x_axis_label
            if y_axis_label:
                chart_obj.y_axis.title = y_axis_label

        elif input_chart_type_upper == CHART_TYPE_LINE:
            chart_obj = LineChart()

            s_min_row, s_min_col, s_max_row, s_max_col = series_data_range
            series_ref = Reference(self.worksheet,
                                   min_col=s_min_col, min_row=s_min_row,
                                   max_col=s_max_col, max_row=s_max_row)
            chart_obj.add_data(series_ref, titles_from_data=titles_from_data)

            if category_labels_range:
                c_min_row, c_min_col, c_max_row, c_max_col = category_labels_range
                cat_ref = Reference(self.worksheet,
                                    min_col=c_min_col, min_row=c_min_row,
                                    max_col=c_max_col, max_row=c_max_row)
                chart_obj.set_categories(cat_ref)

            if title:
                chart_obj.title = title
            if x_axis_label:
                chart_obj.x_axis.title = x_axis_label
            if y_axis_label:
                chart_obj.y_axis.title = y_axis_label

        elif input_chart_type_upper == CHART_TYPE_PIE:
            chart_obj = PieChart()

            s_min_row, s_min_col, s_max_row, s_max_col = series_data_range
            series_ref = Reference(self.worksheet,
                                   min_col=s_min_col, min_row=s_min_row,
                                   max_col=s_max_col, max_row=s_max_row)
            chart_obj.add_data(series_ref, titles_from_data=titles_from_data)

            if category_labels_range:
                c_min_row, c_min_col, c_max_row, c_max_col = category_labels_range
                cat_ref = Reference(self.worksheet,
                                    min_col=c_min_col, min_row=c_min_row,
                                    max_col=c_max_col, max_row=c_max_row)
                chart_obj.set_categories(cat_ref)
            # else:
            #   print("Warning: category_labels_range not provided for PIE chart. Slices may lack labels.")

            if title:
                chart_obj.title = title
            # x_axis_label and y_axis_label are not typically applicable to Pie charts.

        if chart_obj:
            chart_obj.width = chart_width
            chart_obj.height = chart_height
            self.worksheet.add_chart(chart_obj, cell_anchor)
            return chart_obj
        else:
            # This case should not be reached if initial validation for supported types is exhaustive
            # and all supported types have their logic implemented.
            print(f"Warning: Chart type '{chart_type}' passed validation but was not constructed.") # Should use original chart_type for warning
            return None
